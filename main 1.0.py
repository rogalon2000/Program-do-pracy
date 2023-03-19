from mailmerge import MailMerge
from datetime import date
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import openpyxl
import os

from tkinter import filedialog
template = "formularz.docx"
template_ryzykozz = "ryzykozz.docx"
template_upowaznienie = "upowaznienie.docx"
template_ryzykoz = "ryzykoz.docx"
template_zgoda = "zgoda.docx"
class App(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartFrame, PoradniaFrame, IntensywnaFrame):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(StartFrame)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

class BasicFrame(tk.Frame):

    def popup_window(self, sciezka):
        window = tk.Toplevel()

        label = tk.Label(window, text="Co chcesz zrobić?")
        label.pack(fill='x', padx=50, pady=5)

        def otworzFolder():
            os.startfile(os.path.realpath(sciezka))

        button_wroc = tk.Button(window, text="Wróć", command=window.destroy)
        button_wroc.pack(fill='x')
        button_otworz = tk.Button(window, text="Otwórz folder z dokumentami", command=otworzFolder)
        button_otworz.pack(fill='x')
        button_zamknij = tk.Button(window, text="Wyłącz program", command=app.destroy)
        button_zamknij.pack(fill='x')


    def getExcel(self):
        self.import_file_path = ''
        self.import_file_path = filedialog.askopenfilename()
        self.excel_path.set(self.import_file_path)
        return self.import_file_path


class StartFrame(BasicFrame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        dzien_label = tk.Label(self, text="Drukujesz papiery do poradni, czy inne?", font=30)
        dzien_label.pack(pady=10, padx=10)
        button = tk.Button(self, text = 'PORADNIA', fg='white', font=40, bg='green', command=lambda: controller.show_frame(PoradniaFrame))
        button.pack(fill='both', pady=10, padx=10, expand=True,)
        button = tk.Button(self, text = 'INNE',fg='white', font=40, bg='blue', command=lambda: controller.show_frame(IntensywnaFrame))
        button.pack(fill='both', pady=10, padx=10, expand=True)
class PoradniaFrame(BasicFrame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)

        options = {'padx': 5, 'pady': 5}
        browseButton_Excel = tk.Button(self, text='Wybierz listę pacjentów', command=self.getExcel, bg='green', fg='white',
                                       font=('helvetica', 12, 'bold'))
        browseButton_Excel.pack(**options)
        self.excel_path = tk.StringVar(self)
        self.excel_path.set('Ścieżka do pliku')
        self.excel_path_label = tk.Label(self, textvariable=self.excel_path)
        self.excel_path_label.pack()
        self.dzien_label = tk.Label(self, text="Data poradni:")
        self.dzien_label.pack(**options)
        self.dzien_box = tk.Entry(self)
        self.dzien_box.pack(**options)
        self.lekarz_label = tk.Label(self, text="Lekarz")
        self.lekarz_label.pack(**options)
        self.lekarz_box = tk.Entry(self)
        self.lekarz_box.pack(**options)
        self.submitButton = tk.Button(self, text='Generuj dokumenty', command=self.submit, bg='green', fg='white',
                                       font=('helvetica', 12, 'bold'))
        self.submitButton.pack(**options)
        self.homeframe = tk.Button(self, text = 'home', command=lambda: controller.show_frame(StartFrame))
        self.homeframe.pack(**options)

    def submit(self):
        dzien = self.dzien_box.get()
        lekarz = self.lekarz_box.get()
        self.sciezka = filedialog.askdirectory()
        self.prepare_docs(lekarz, dzien, self.sciezka)

    def prepare_docs(self, lekarz, dzien, folder_docelowy):
        try:
            ps = openpyxl.load_workbook(self.import_file_path)
        except: messagebox.showerror('Error', 'Wybierz plik z listą pacjentów')
        else:
            sheet = ps["Lista pacjentów gabinetu"]
            for row in range(5, sheet.max_row + 1):
                document = MailMerge(template)
                imie = str(sheet['D' + str(row)].value)
                nazwisko = str(sheet['C' + str(row)].value)
                PESEL = str(sheet['E' + str(row)].value)
                document.merge(Wypełniający=lekarz,
                               Data_zgody=dzien,
                               Pesel_data_urodzenia=PESEL,
                               Nazwisko_Imię=nazwisko + ' ' + imie)
                document.write(str(folder_docelowy)+'\ '+nazwisko + ' ' + imie + '.docx')
            self.popup_window(folder_docelowy)

class IntensywnaFrame(BasicFrame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        options = {'padx': 5, 'pady': 5}
        title_label = tk.Label(self, text="Inne dokumenty (przyjęcie, blok)")
        title_label.pack()
        '''
        Do wyboru: wprowadzenie listy pacjentów lub wprowadzenie danych pojedynczego pacjenta.
        Na razie: generowanie wybranych w checkbuttonach dokumentów dla wszystkich pacjentów z listy
        Docelowo: wygenerowanie checklisty pacjentów z zaimportowanego excela i generowanie wybranych dokumentów tylko dla wybranych pacjentów
        '''

        browseButton_Excel = tk.Button(self, text='Wybierz listę pacjentów', command=self.getExcel, bg='green', fg='white',
                                       font=('helvetica', 12, 'bold'))

        def show_hide(rbvariable):
            if rbvariable.get()==1:
                browseButton_Excel.pack(side='top', after=self.wybor2)
                self.excel_path_label.pack(side='top', after=browseButton_Excel)
                self.pesel_box.pack_forget()
                self.imie_box.pack_forget()
                self.nazwisko_box.pack_forget()
                self.pesel_label.pack_forget()
                self.imie_label.pack_forget()
                self.nazwisko_label.pack_forget()
            elif rbvariable.get()==2:
                browseButton_Excel.pack_forget()
                self.excel_path_label.pack_forget()
                self.nazwisko_box.pack(side='top', after=self.wybor2)
                self.nazwisko_label.pack(side='top', after=self.wybor2)
                self.imie_box.pack(side='top', after=self.wybor2)
                self.imie_label.pack(side='top', after=self.wybor2)
                self.pesel_box.pack(side='top', after=self.wybor2)
                self.pesel_label.pack(side='top', after=self.wybor2)

        self.v = tk.IntVar()
        self.wybor = tk.Radiobutton(self, anchor='n', text = 'Lista pacjentów', variable=self.v, value=1, command=lambda: show_hide(self.v))
        self.wybor2 = tk.Radiobutton(self, anchor='n', text = 'Pojedynczy pacjent', variable=self.v, value=2, command=lambda: show_hide(self.v))
        self.wybor.pack()
        self.wybor2.pack()

### wyświetlanie ścieżki do listy pacjentów
        self.excel_path = tk.StringVar(self)
        self.excel_path.set('')
        self.excel_path_label = tk.Label(self, textvariable=self.excel_path)
### panel do wpisywania danych pacjenta pojedynczego
        self.pesel_label = tk.Label(self, text='PESEL')
        self.pesel_box = tk.Entry(self)
        self.imie_label = tk.Label(self, text='Imię')
        self.imie_box = tk.Entry(self)
        self.nazwisko_label = tk.Label(self, text='Nazwisko')
        self.nazwisko_box = tk.Entry(self)

### box do podania daty i danych lekarza
        self.dzien_label = tk.Label(self, text='Data')
        self.dzien_box = tk.Entry(self)
        self.lekarz_label = tk.Label(self, text='Lekarz')
        self.lekarz_box = tk.Entry(self)
        self.dzien_label.pack()
        self.dzien_box.pack()
        self.lekarz_label.pack()
        self.lekarz_box.pack()
### zmienne do odznaczania, które dokumenty mają być drukowane
        self.checkvar1 = tk.IntVar()
        self.checkvar1.set(0)
        self.checkvar2 = tk.IntVar()
        self.checkvar2.set(0)
        self.checkvar3 = tk.IntVar()
        self.checkvar3.set(0)
        self.checkvar4 = tk.IntVar()
        self.checkvar4.set(0)
        self.checkvar5 = tk.IntVar()
        self.checkvar5.set(0)

### Lista Labeli z dokumentami, które mają być wydrukowane
        self.document1 = tk.Label(self, text = 'Upoważnienie do informacji medycznej')
        self.document2 = tk.Label(self, text = 'Ryzyko zakrzepowo-zatorowe')
        self.document3 = tk.Label(self, text = 'Ryzyko zakażenia')
        self.document4 = tk.Label(self, text = 'Zgoda na hospitalizację')
        self.document5 = tk.Label(self, text = '"Pacjent nieprzytomny, nie jest w stanie złożyć podpisu"')

### Przycisk do otwarcia nowego okna, w którym będą wybierane dokumenty:







### Przyciski do odhaczania, które dokumenty mają być drukowane - mają być docelowo w nowym oknie
        '''
        
        
        do przerobienia
        
        
        '''
        doc1 = tk.Checkbutton(self, text = 'Upoważnienie do informacji medycznej', variable=self.checkvar1)
        doc1.pack(anchor='w')

        doc2 = tk.Checkbutton(self, text = 'Ryzyko zakrzepowo-zatorowe', variable=self.checkvar2)
        doc2.pack(anchor='w')

        doc3 = tk.Checkbutton(self, text = 'Ryzyko zakażenia', variable=self.checkvar3)
        doc3.pack(anchor='w')

        doc4 = tk.Checkbutton(self, text = 'Zgoda na hospitalizację', variable=self.checkvar4)
        doc4.pack(anchor='w')

        doc5 = tk.Checkbutton(self, text = 'Pacjent nieprzytomny?', variable=self.checkvar5)
        doc5.pack(anchor='w')
### przycisk do generowania dokumentów

        self.submitButton = tk.Button(self, text='Generuj dokumenty', command=self.submit, bg='green', fg='white',
                                   font=('helvetica', 12, 'bold'))
        self.submitButton.pack()
### przycisk wróć
        homeframe = tk.Button(self, text = 'Powrót',fg='white', bg='red', font=30, command=lambda: controller.show_frame(StartFrame))
        homeframe.pack(ipady=15, ipadx=15, expand=True)
### funkcje do generowania dokumentów na OIT
    def submit(self):
        dzien = self.dzien_box.get()
        lekarz = self.lekarz_box.get()
        imie = self.imie_box.get()
        nazwisko = self.nazwisko_box.get()
        PESEL = self.pesel_box.get()
        if self.v.get() == 1:
            messagebox.showinfo('Info', 'Ta funkcja nie jest jeszcze zaimplementowana')
            ###self.sciezka = filedialog.askdirectory()
            ###self.prepare_docs(lekarz, dzien, imie, nazwisko, PESEL, self.checkvar5.get(), self.sciezka)
        elif self.v.get() == 2:
            self.sciezka = filedialog.askdirectory()
            self.prepare_docs(lekarz, dzien, imie, nazwisko, PESEL, self.checkvar5.get(), self.sciezka)

    def prepare_docs(self, lekarz, dzien, imie, nazwisko, PESEL, nieprzytomny, folder_docelowy):
        lista_dokumentow = []
        if self.checkvar1.get()==1:
            lista_dokumentow.append(template_upowaznienie)
        if self.checkvar2.get()==1:
            lista_dokumentow.append(template_ryzykozz)
        if self.checkvar3.get()==1:
            lista_dokumentow.append(template_ryzykoz)
        if self.checkvar4.get() == 1:
            lista_dokumentow.append(template_zgoda)
        print(lista_dokumentow)
        if self.v.get() == 1:
            try:
                ps = openpyxl.load_workbook(self.import_file_path)
            except:
                messagebox.showerror('Error', 'Wybierz plik z listą pacjentów')
            else:
                pass
        elif self.v.get() == 2:
            try:
                False ### do wymyślenia jak ma sprawdzać, czy uzupełnione są dane pacjenta
            except: messagebox.showerror('Error', 'Uzupełnij dane pacjenta')
            else:
                for dokument in lista_dokumentow:
                    docelowy = MailMerge(dokument)
                    docelowy.merge(lekarz=lekarz,
                               data=dzien,
                               pesel=PESEL,
                               nazwisko=nazwisko + ' ' + imie)
                    if nieprzytomny == 1:
                        docelowy.merge(nieprzytomny='Pacjent nieprzytomny, nie jest w stanie złożyć podpisu.')
                    docelowy.write(str(folder_docelowy)+'\ '+nazwisko + ' ' + imie + ' ' + str(dokument))
                self.popup_window(folder_docelowy)


if __name__ == "__main__":
    app = App()
    app.geometry('800x600')
    app.mainloop()

